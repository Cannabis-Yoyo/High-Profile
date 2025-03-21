import time
import random
import re
import pandas as pd
import nltk
from fractions import Fraction
from nltk.stem import WordNetLemmatizer
from openpyxl import load_workbook
from selenium import webdriver
from rapidfuzz import fuzz
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options  # Use Firefox options instead of Chrome options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import streamlit as st
import sys
import os
import requests
import io
from streamlit_lottie import st_lottie
import shutil
from webdriver_manager.firefox import GeckoDriverManager  # Use GeckoDriverManager for Firefox
from selenium.webdriver.firefox.service import Service  # Service for Firefox WebDriver


# import time
# import random
# import re
# import pandas as pd
# import nltk
# from fractions import Fraction
# from nltk.stem import WordNetLemmatizer
# from openpyxl import load_workbook
# from selenium import webdriver
# from rapidfuzz import fuzz
# from selenium.webdriver.common.by import By
# # from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# # import undetected_chromedriver as uc
# import streamlit as st
# import sys
# import os
# import requests
# import io
# from streamlit_lottie import st_lottie
# import shutil

# # from webdriver_manager.chrome import ChromeDriverManager



nltk.download('wordnet')
lemmatizer = WordNetLemmatizer()

st.markdown('<style>div.block-container{padding-top:2rem;}</style>', unsafe_allow_html=True)

def load_lottie_url(url: str):
    import requests
    r = requests.get(url)
    if r.status_code != 200:
        return None
    return r.json()

# Custom CSS to center only the title
st.markdown("""
    <style>
        /* Center the title */
        .title {
            text-align: center;
            color: #00000;
            font-size: 32px;
        }

        /* Style the main page background */
        
        .stApp {
            background-color: #F4F7FC; /* Light Gray-Blue */
        }
        
        .title-container h1 {
            color: white !important; /* Force White Text */
            font-size: 32px;
            font-weight: bold;
        }

        /* Style the header background */
        .stMarkdown h1 {
            background-color: #6a2af1; /* Dark Blue */
            color: ffff;
            padding: 15px;
            border-radius: 8px;
            text-align: center;
        }

        /* Style the sidebar */
        .css-1d391kg {
            background-color: #EAECEF; /* Light Gray */
            padding: 20px;
            border-radius: 10px;
        }

        /* Buttons */
        .stButton>button {
            background-color: #1E3A8A; /* Dark Blue */
            color: white;
            border-radius: 5px;
            padding: 10px;
            font-weight: bold;
            border: none;
            transition: 0.3s;
        }

        .stButton>button:hover {
            background-color: #1E40AF; /* Slightly Brighter Blue */
        }

        /* Style file uploader */
        .stFileUploader {
            background-color: white;
            border-radius: 8px;
            padding: 10px;
            border: 1px solid #B0BEC5; /* Gray Border */
        }

        /* Style the download button */
        .stDownloadButton>button {
            background-color: #10B981; /* Green */
            color: white;
            border-radius: 5px;
            padding: 10px;
            font-weight: bold;
            border: none;
            transition: 0.3s;
        }

        .stDownloadButton>button:hover {
            background-color: #059669; /* Darker Green */
        }
    </style>
""", unsafe_allow_html=True)

# Title at the top (centered)
# st.markdown("<h1 class='title'>Web Scraper for High-Profile Cannabis Data</h1>", unsafe_allow_html=True)

st.markdown("""
    <div class="title-container">
        <h1 class="title-text">Web Scraper for High-Profile Cannabis Data</h1>
    </div>
""", unsafe_allow_html=True)

# Layout with 2 columns: Content + Animation
col1, col2 = st.columns([2, 1])  # Adjust ratio for better space utilization


# LEFT SIDE: Main Scraper UI
with col1:
    # Add instructions in the sidebar
    st.sidebar.header("Instructions")
    st.sidebar.write("""
        1. **Upload an Excel file** that contains product pricing data for scraping.
        2. Click **Start Scraping** to begin the extraction process.
        3. The scraper will extract product details and update the file.
        4. After the scraping process, you will be able to download the updated file.
        5. The scraper will process different product categories.
    """)

# Now, place the file uploader below the instructions
uploaded_file = st.sidebar.file_uploader("Upload Excel File", type=["xlsx"])


if uploaded_file:
    # Load the file after it's uploaded
    st.sidebar.write(f"📂 Uploaded File: **{uploaded_file.name}**")
    df = pd.read_excel(uploaded_file, sheet_name="Pricing Research")
    df = df.astype(str)

    # Button to start scraping
    if st.sidebar.button("🚀 Start Scraping"):
        st.write("🔄 **Processing... Please wait!**")
        
    # DataFrame modifications and scraping logic will follow as before...
    df['Weight'] = df['Weight'].astype(str).str.strip()

    # # Now you can perform operations on the DataFrame
    # df['Weight'] = df['Weight'].astype(str).str.strip()

    # Mapping dictionary to standardize category names
    category_mapping = {
        "Preroll": "Pre-Rolls",
        "Preroll Infused": "Pre-Rolls",
        "Singles": "Pre-Rolls",
        "Infused Pre-Rolls": "Pre-Rolls",
        "Shake": "Flower",
        "Trim": "Flower",
        "Vape Pens": "Vape",
        "Drinks": "Edibles",
        "Chocolates": "Edibles",
        "Gummies": "Edibles",
        "Baked Good": "Edibles",
        "Capsules": "Edibles",
        "Tablets": "Edibles",
        "Cartridges": "Vape",
        "Disposables": "Vape",
        "Live Resin": "Concentrates",
        "Wax": "Concentrates",
        "Rosin": "Concentrates",
        "Crumble": "Concentrates",
        "Batteries": "Accessories",
        "Lighters": "Accessories",
        "Papers": "Accessories",
        "Rolling Supplies": "Accessories",
        "Gift Cards": "Accessories",
        "Glassware": "Accessories",
        "Dab Tools": "Accessories",
        "Cleaning Solution": "Accessories",
        "Storage": "Accessories",
        "Containers": "Accessories",
        "Topicals": "Topicals"
    }
    df['Category'] = df['Category'].replace(category_mapping)

    # Mapping dictionary to standardize brand names
    brand_mapping = {
        "&Shine": "&Shine",
        "1906": "1906",
        "93 Boyz": "93 Boyz",
        "Aeriz": "Aeriz",
        "Bloom": "Bloom",
        "CANNECT WELLNESS": "Cannect Wellness",
        "Cookies": "Cookies",
        "Cresco": "Cresco",
        "Daze off": "Daze Off",
        "Doctor Solomon's": "Doctor Solomon's",
        "Dogwalkers": "Dogwalkers",
        "Effin' Edibles": "Effin'",
        "Fig Farms": "Fig Farms",
        "Floracal": "FloraCal Farms",
        "Good Green": "Good Green",
        "Good News": "Good News",
        "Good Tide": "Good Tide",
        "High Supply": "High Supply",
        "Incredibles": "Incredibles",
        "Joss": "Joos",
        "Journeyman": "Journeyman",
        "Kiva Confections": "Kiva",
        "Legacy Cannabis  LLC": "Legacy",
        "Mindy's": "Mindy's Edibles",
        "Nature's Grace and Wellness": "Nature's Grace and Wellness",
        "Ozone": "Ozone",
        "Paul Bunyan": "Paul Bunyan",
        "Revolution": "Revolution Cannabis",
        "RYTHM": "Rythm",
        "Simply Herb": "Simply Herb",
        "Spring Lake": "Spring Lake",
        "Superflux": "Superflux",
        "Tales and Travels": "Tales & Travels",
        "The Botanist": "The Botanist",
        "The Funnies": "The Funnies",
        "Tonic": "Tonic",
        "Uncle Arnie's": "Uncle Arnie's",
        "UpNorth Humbolt": "UpNorth",
        "Verano": "Verano",
        "Wonder Wellness": "Wonder",
        "Wyld": "Wyld"
    }
    df['Brand'] = df['Brand'].replace(brand_mapping)

    unique_categories = df['Category'].dropna().unique()
    print("Unique Categories from Excel:", unique_categories)

    # -------------------------------
    # Category -> URL Map
    # -------------------------------
    category_url_map = {
        "Pre-Rolls": "https://highprofilecannabis.com/shop/martinsville-dispensary/pre-rolls",
        "Flower": "https://highprofilecannabis.com/shop/martinsville-dispensary/flower",
        "Edibles": "https://highprofilecannabis.com/shop/martinsville-dispensary/edibles",
        "Vape": "https://highprofilecannabis.com/shop/martinsville-dispensary/vape",
        "Tincture & Capsules": "https://highprofilecannabis.com/shop/martinsville-dispensary/tinctures",
        "Accessories": "https://highprofilecannabis.com/shop/martinsville-dispensary/accessories",
        "Concentrates": "https://highprofilecannabis.com/shop/martinsville-dispensary/concentrates",
        "Topicals": "https://highprofilecannabis.com/shop/martinsville-dispensary/topicals"
    }

    # user_agent = (
    #     "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    #     "AppleWebKit/537.36 (KHTML, like Gecko) "
    #     "Chrome/122.0.0.0 Safari/537.36"
    # )

    user_agents = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    ]
    
    from selenium import webdriver
    from selenium.webdriver.firefox.options import Options
    from webdriver_manager.firefox import GeckoDriverManager
    from selenium.webdriver.firefox.service import Service
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    
    # Set up Firefox options for headless browsing
    options = Options()
    options.add_argument("--headless")  # Run Firefox in headless mode
    options.add_argument("--disable-gpu")  # Optional: Disable GPU for headless mode
    
    # Function to get the Firefox driver
    @st.cache_resource
    def get_driver():
        # Automatically install the correct version of GeckoDriver
        driver = webdriver.Firefox(
            service=Service(GeckoDriverManager().install()),  # Ensure the correct GeckoDriver is installed
            options=options
        )
        return driver

    # Initialize the Firefox driver using the updated method
    driver = get_driver()
    
    # Example to load a page using Firefox driver
    try:
        driver.get("https://highprofilecannabis.com")
        
        # Wait for the body of the page to load (adjust as needed)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, "//body"))
        )
        
        st.code(driver.page_source)  # Display page source for debugging
    
    except Exception as e:
        st.error(f"Error loading page: {e}")
        driver.quit()
    
    # Close the browser session after the task
    driver.quit()

    
    # # Initialize the WebDriver using the updated method
    # driver = get_driver()
    
    # # Example test to load a page
    # try:
    #     driver.get("https://highprofilecannabis.com")
        
    #     # Wait for the body of the page to load (adjust as needed)
    #     WebDriverWait(driver, 30).until(
    #         EC.presence_of_element_located((By.XPATH, "//body"))
    #     )
        
    #     st.code(driver.page_source)  # Display page source for debugging
    
    # except Exception as e:
    #     st.error(f"Error loading page: {e}")
    #     driver.quit()
    
    # # Close the browser session after the task
    # driver.quit()
    

    
    # from selenium import webdriver
    # from selenium.webdriver.chrome.options import Options
    # from selenium.webdriver.chrome.service import Service
    # from webdriver_manager.chrome import ChromeDriverManager
    # from webdriver_manager.core.os_manager import ChromeType

    # @st.cache_resource
    # def get_driver():
    #     return webdriver.Chrome(
    #         service=Service(
    #             ChromeDriverManager(chrome_type=ChromeType.CHROMIUM).install()
    #         ),
    #         options=options,
    #     )

    # # options = Options()
    # options = uc.ChromeOptions()
    # options.add_argument("--disable-gpu")
    # options.add_argument("--headless")
    # driver = uc.Chrome(options=options)

    # driver = get_driver()
    # st.code(driver.page_source)


    
    # from selenium import webdriver
    # from selenium.webdriver.chrome.options import Options
    # from selenium.webdriver.chrome.service import Service
    # from webdriver_manager.chrome import ChromeDriverManager
    # from webdriver_manager.core.os_manager import ChromeType
    
    # @st.cache_resource
    # def get_driver():
    #     options = Options()
    #     options.add_argument("--disable-gpu")
    #     options.add_argument("--headless")  # Ensures headless mode
        
    #     return webdriver.Chrome(
    #         service=Service(ChromeDriverManager(chrome_type=ChromeType.CHROMIUM).install()),
    #         options=options,
    #     )

    # # Specify the exact version of Chromium that you're using
    # chromedriver_path = ChromeDriverManager(version="114.0.5735.90").install()
    
    # # Create a Chrome driver with the specified version
    # driver = webdriver.Chrome(executable_path=chromedriver_path, options=options)
    
    # def get_driver():
    #     # Auto-install the correct ChromeDriver version
    #     chromedriver_autoinstaller.install()
    
    #     # Set Chrome options
    #     chrome_options = Options()
    #     chrome_options.add_argument("--headless")  # Run headless mode
    #     chrome_options.add_argument("--no-sandbox")
    #     chrome_options.add_argument("--disable-dev-shm-usage")
    #     chrome_options.add_argument("--disable-gpu")
    #     chrome_options.add_argument("--window-size=1920,1080")
    
    #     try:
    #         # Try launching Chrome
    #         driver = webdriver.Chrome(options=chrome_options)
    #         return driver
    #     except Exception as e:
    #         raise Exception(f"Chrome launch failed: {e}")
    
    
    # def get_driver():
    #     # Ensure Chromium and Chromedriver exist in deployment environment
    #     chrome_path = shutil.which("google-chrome") or shutil.which("chromium-browser") or shutil.which("chromium")
    #     chromedriver_path = shutil.which("chromedriver")
    
    #     if not chrome_path or not chromedriver_path:
    #         raise Exception("Chrome or Chromedriver not found in the system!")
    
    #     # Set Chrome options
    #     chrome_options = Options()
    #     chrome_options.add_argument("--headless")  # Run Chrome in headless mode
    #     chrome_options.add_argument("--no-sandbox")
    #     chrome_options.add_argument("--disable-dev-shm-usage")
    #     chrome_options.add_argument("--disable-gpu")
    #     chrome_options.add_argument("--window-size=1920,1080")
        
    #     # Configure WebDriver service
    #     service = Service(chromedriver_path)
    #     driver = webdriver.Chrome(service=service, options=chrome_options)
    
    #     return driver

    
    # def get_driver():
    #     chrome_options = Options()
    #     # For debugging: Remove headless option to run with GUI temporarily
    #     # chrome_options.add_argument("--headless")  # You can comment this out to run in non-headless mode
    #     chrome_options.add_argument("--disable-gpu")  # Disable GPU hardware acceleration
    #     chrome_options.add_argument("--no-sandbox")  # Run without sandbox (important for some environments)
    #     chrome_options.add_argument("window-size=1920x1080")  # Full screen to avoid rendering issues

    #     driver = webdriver.Chrome(options=chrome_options)
    #     driver.set_page_load_timeout(600)  # Set the page load timeout to 10 minutes
    #     return driver
    
    def wait_for_page_load(driver, xpath, timeout=60):
        """Wait for a specific element to load on the page."""
        try:
            # Wait for the element to appear on the page, indicating that it is fully loaded.
            WebDriverWait(driver, timeout).until(
                EC.visibility_of_element_located((By.XPATH, xpath))
            )
            print("Page loaded successfully!")
        except Exception as e:
            print(f"Error waiting for page to load: {e}")
            driver.quit()  # Close the driver after an error
            raise e

    response = requests.get('https://api.ipify.org?format=json')
    print(f"Current IP: {response.json()['ip']}")

    def click_element(driver, xpath, description, wait_time=5):
        """Scrolls to element, ensures visibility, and clicks it.
        If normal click fails, attempts JavaScript click."""
        try:
            element = WebDriverWait(driver, wait_time).until(
                EC.element_to_be_clickable((By.XPATH, xpath))
            )
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
            time.sleep(1)  # Allow animations
            element.click()
            print(f"✅ Clicked: {description}")
            time.sleep(3)
            return True
        except Exception:
            try:
                element = driver.find_element(By.XPATH, xpath)
                driver.execute_script("arguments[0].click();", element)
                time.sleep(3)
                return True
            except Exception as js_e:
                return False

    def normalize_weight(weight):
        weight_str = str(weight).lower().strip().replace(" ", "")
        if weight_str in ["ea", "each", "1ea", "1each"]:
            return "SKIP_WEIGHT"
        if weight_str in ["1/8thoz", "1/8th"]:
            return "1/8oz"
        if "mg" in weight_str:
            try:
                mg_value = float(weight_str.replace("mg", ""))
                grams = mg_value / 1000  # Convert to grams
                if grams < 1:
                    return f".{str(grams)[2:]}g"  # '.75g' instead of '0.75g'
                else:
                    return f"{grams:g}g"
            except Exception:
                return None
        elif "g" in weight_str:
            try:
                g_value = float(weight_str.replace("g", ""))
                if g_value < 1 and g_value > 0:
                    return f".{str(g_value)[2:]}g"
                else:
                    return f"{g_value:g}g"
            except Exception:
                return None
        elif "oz" in weight_str:
            return weight_str  # Keep oz values unchanged
        else:
            return weight_str

    def extract_weight_from_name(product_name):
        """
        Extracts weight information from product names using regex and removes it.
        """
        special_weight_pattern = r'\(\s*\d+(?:\.\d+)?\s*(?:g|mg|oz)\s*(ea|total).*\)' 
        special_match = re.search(special_weight_pattern, product_name, re.IGNORECASE)

        if special_match:
            extracted_weight = special_match.group(0)
            clean_name = product_name
            return extracted_weight, clean_name

        pattern = r'(\d+(?:\.\d+)?|\d+/\d+)\s*(mg|g|oz)'
        match = re.search(pattern, product_name, re.IGNORECASE)

        if not match:
            return None, product_name

        numeric_part = match.group(1).lower()
        unit_part = match.group(2).lower()

        if '/' in numeric_part and unit_part == 'oz':
            extracted_weight = numeric_part + unit_part
        else:
            try:
                numeric_value = float(numeric_part)
                if unit_part == 'mg':
                    g_value = numeric_value / 1000.0
                    extracted_weight = f"{g_value:g}g"
                elif unit_part == 'g':
                    extracted_weight = f"{numeric_value:g}g"
                elif unit_part == 'oz':
                    extracted_weight = f"{numeric_value:g}oz"
            except ValueError:
                return None, product_name

        clean_name = re.sub(pattern, '', product_name, flags=re.IGNORECASE).strip()

        return extracted_weight, clean_name

    def convert_weight_alternate(weight_str):
        weight_str = weight_str.lower().replace("th", "").strip()
        if "g" in weight_str:
            try:
                grams = float(weight_str.replace("g", "").strip())
                ounces = grams / 28.3495
                fractional_oz = Fraction(ounces).limit_denominator(8)
                return f"{fractional_oz}oz"
            except Exception:
                return None
        elif "oz" in weight_str:
            try:
                oz_value = weight_str.replace("oz", "").strip()
                if "/" in oz_value:
                    ounces = float(Fraction(oz_value))
                else:
                    ounces = float(oz_value)
                grams = ounces * 28.3495
                grams_str = f"{grams:.1f}g"
                if grams_str.endswith(".0g"):
                    grams_str = grams_str.replace(".0g", "g")
                return grams_str
            except Exception:
                return None
        else:
            return None

    def update_weight_in_excel(row_index, new_weight):
        try:
            wb = load_workbook(uploaded_file)
            ws = wb["Pricing Research"]
            weight_col = "Q"
            row_pos = row_index + 2
            ws[f"{weight_col}{row_pos}"] = new_weight
            wb.save(uploaded_file)
            print(f"Updated Excel row {row_pos} with new weight: {new_weight}/n")
        except Exception:
            print(f"Weight not updated for row {row_index}")

    def extract_data_and_update_excel(row_index, price, thc, url):
        try:
            wb = load_workbook(uploaded_file)
            ws = wb["Pricing Research"]
            price_col = "AG"
            thc_col = "AH"
            url_col = "AI"
            row_pos = row_index + 2
            
            if price is not None:
                ws[f"{price_col}{row_pos}"] = str(price) if isinstance(price, (str, int, float)) else ""
            if thc is not None:
                ws[f"{thc_col}{row_pos}"] = str(thc) if isinstance(thc, (str, int, float)) else ""
            if url is not None:
                ws[f"{url_col}{row_pos}"] = str(url) if isinstance(url, (str, int, float)) else ""

            wb.save(uploaded_file)
            st.success(f"Excel updated successfully for Row {row_pos}!")
        except Exception as e:
            st.error(f"Failed to update Excel: {e}")

    def extract_keywords(text):
        text = text.lower()
        text = re.sub(r'[^a-z0-9\s]', '', text)
        tokens = text.split()
        stopwords = {"essentials", "all", "in", "one", "select", "disposable", "vape", "the", "a", "an", "of", "and"}
        filtered_tokens = [token for token in tokens if token not in stopwords]
        return filtered_tokens

    def lemmatize_keywords(keywords):
        return [lemmatizer.lemmatize(word) for word in keywords]

    # def confirm_age(driver):
    #     try:
    #         age_button = WebDriverWait(driver, 10).until(
    #             EC.element_to_be_clickable((By.ID, "age-gate-yes"))
    #         )
    #         age_button.click()
    #         time.sleep(2)
    #     except Exception:
    #         print("Age Confirmation not found. Please Run the Code Again!")
            
            
    # def confirm_age(driver):
    #     try:
    #         # Wait for the age modal "Yes" button to be clickable.
    #         age_button = WebDriverWait(driver, 10).until(
    #             EC.element_to_be_clickable((By.ID, "age-gate-yes"))
    #         )
    #         age_button.click()
    #         print("--------------")
    #         print("Age confirmed!")
    #         print("--------------")
    #         time.sleep(2)  # Wait a bit after clicking.
    #     except Exception:
    #         print("Age Confirmation not found. Please Run the Code Again!")

    def confirm_age(driver):
        try:
            age_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "age-gate-yes"))
            )
            age_button.click()
            print("Age confirmed!")
            time.sleep(2)
        except Exception:
            print("Age confirmation not found.")

    def clear_filters(driver):
        try:
            selected_checkboxes = driver.find_elements(By.CSS_SELECTOR, "input[type='checkbox']:checked")
            for checkbox in selected_checkboxes:
                driver.execute_script("arguments[0].click();", checkbox)
                time.sleep(1)
            print("Manually Cleared Selected Brand Filter.")
        except Exception:
            print("Brand Not Found on Website")

    def clear_weight_filters(driver):
        weights_xpath = '//p[contains(@class, "shoppage__filters-option") and text()="Weights"]'
        if not click_element(driver, weights_xpath, "Weights Filter"):
            print("Could not open the Weights panel to clear filters.")
            return

        time.sleep(2)

        try:
            selected_spans = driver.find_elements(By.XPATH, '//span[contains(@class, "shoppage__filters-weights-flex-item-selected")]')

            if not selected_spans:
                print("No weight filters appear to be selected.")
                return

            for span in selected_spans:
                driver.execute_script("arguments[0].click();", span)
                time.sleep(1)
            print("Cleared all active weight filters by clicking selected spans.")
        except Exception:
            print(f"Weight Filter Uncleared")

    def search_brand(driver, brand):
        brands_xpath = '//p[contains(@class, "shoppage__filters-option") and text()="Brands"]'
        if click_element(driver, brands_xpath, "Brands Filter"):
            xpath_brand = f"//label[contains(text(), '{brand}')]"
            if click_element(driver, xpath_brand, f"Brand: {brand}"):
                time.sleep(2)
                return True
        print(f"Brand {brand} not found")
        return False

    def attempt_weight_selection(driver, weight):
        if not weight:
            return False

        weights_xpath = '//p[contains(@class, "shoppage__filters-option") and text()="Weights"]'
        weight_xpath = f'//div[contains(@class, "shoppage__filters-weights-item")]/span[text()="{weight}"]'

        if click_element(driver, weights_xpath, "Weights Filter"):
            if click_element(driver, weight_xpath, f"Weight: {weight}"):
                time.sleep(5)
                return True

        return False

    def add_leading_zero_variants(weight_str):
        variants = []
        if not weight_str:
            return variants
        
        variants.append(weight_str)
        if weight_str.startswith('.'):
            zero_variant = '0' + weight_str
            variants.append(zero_variant)
        return variants

    def select_weight(driver, weight, category, product_name):
        if not weight:
            return False 

        normalized_weight = normalize_weight(weight)
        alternate_weight = convert_weight_alternate(normalized_weight)
        extracted_weight, _ = extract_weight_from_name(product_name)

        weight_options = []
        weight_options.extend(add_leading_zero_variants(normalized_weight))
        if alternate_weight:
            weight_options.extend(add_leading_zero_variants(alternate_weight))
        if extracted_weight:
            weight_options.extend(add_leading_zero_variants(extracted_weight))

        weight_options = list(dict.fromkeys(weight_options))

        for w in weight_options:
            if attempt_weight_selection(driver, w):
                return True

        return False

    def match_product(driver, expected_product_name, brand_name, weight, weight_applied, expected_quantity):
        """
        Matches products by comparing keyword similarity.
        
        - Shows **all products being compared** along with their scores.
        - If more than one product achieves the highest score (≥ 50% threshold),
        returns a list of all those products.
        """
        product_list_xpath = '//div[contains(@class, "shopitem")]'
        try:
            products = WebDriverWait(driver, 15).until(
                EC.presence_of_all_elements_located((By.XPATH, product_list_xpath))
            )
        except Exception as e:
            print("No product found!")
            return None

        # Remove the brand name from the expected product name before comparison.
        expected_product_name_cleaned = expected_product_name.replace(brand_name, "").strip()

        # If weight was applied, remove weight from the product name.
        if weight_applied:
            expected_product_name_cleaned = expected_product_name_cleaned.replace(weight, "").strip()
            print(f"Weight '{weight}' was applied. Removed weight from product name: '{expected_product_name_cleaned}'")

        expected_keywords = lemmatize_keywords(extract_keywords(expected_product_name_cleaned))
        print(f"Extracted Keywords from Expected Product Name: {expected_keywords}")

        best_matches = []
        highest_score = 0
        unique_products = set()  # Track unique product names to avoid duplicates

        # Prepare the cleaned text for fuzzy matching
        expected_clean = expected_product_name_cleaned.lower()
        token_count = len(expected_clean.split())

        # Set a dynamic fuzzy matching threshold based on the number of words
        fuzzy_threshold = 60 if token_count < 3 else 75

        print("\nComparing Products:\n")

        for product in products:
            try:
                product_name_element = product.find_element(By.CLASS_NAME, "shopitem__title")
                product_name = product_name_element.text.strip() if product_name_element else None

                if not product_name or product_name in unique_products:
                    continue

                unique_products.add(product_name)

                # NEW: Extract candidate quantity from the product name.
                candidate_quantity, _ = extract_quantity_from_name(product_name)

                # NEW: Quantity matching logic
                eq_lower = expected_quantity.lower() if expected_quantity else None
                cq_lower = candidate_quantity.lower() if candidate_quantity else None

                if eq_lower and cq_lower and eq_lower != cq_lower:
                    print(f"Both sides have a quantity but they do not match: '{candidate_quantity}' vs '{expected_quantity}'")
                    continue
                else:
                    print("Quantity check passed or not applicable.")

                website_clean = product_name.lower()
                fuzzy_score = fuzz.token_set_ratio(expected_clean, website_clean)
                print(f"Comparing with: '{product_name}' | Fuzzy Score: {fuzzy_score}")

                if fuzzy_score >= fuzzy_threshold:
                    if fuzzy_score > highest_score:
                        highest_score = fuzzy_score
                        best_matches = [{"product_name": product_name, "score": fuzzy_score, "element": product}]
                    elif fuzzy_score == highest_score:
                        best_matches.append({"product_name": product_name, "score": fuzzy_score, "element": product})

            except Exception as e:
                continue

        if highest_score < fuzzy_threshold:
            print(f"Best fuzzy score ({highest_score}) is below threshold ({fuzzy_threshold}) → Product UNMATCHED.")
            return None

        print("\nBest Matches (Fuzzy Score above threshold):")
        for match in best_matches:
            print(f" - {match['product_name']} (Fuzzy Score: {match['score']})")

        return best_matches

    def extract_quantity_from_name(product_name):
        """
        Extracts quantity information from product names using regex and removes it.
        """

        # Regular expression pattern to match quantities like "20pk", "10ct", "2ct"
        pattern = r'(\d+)(?:\s*)(pk|ct)\b'  # Matches "20pk", "10ct", "2ct"

        match = re.search(pattern, product_name, re.IGNORECASE)

        if match:
            extracted_quantity = match.group(0).strip()  # Extract the full quantity (e.g., "20pk")
            clean_name = re.sub(pattern, '', product_name, flags=re.IGNORECASE).strip()  # Remove it from the product name
            return extracted_quantity, clean_name

        return None, product_name  # If no quantity found, return None

    def extract_ratio_from_name(product_name):
        """
        Extracts the ratio from product names using regex and removes it.
        """
        pattern = r'\(?([A-Z:]+\s\d+:\d+(?::\d+)*)\)?'  # Regex pattern to match ratios like "THC:CBD 1:1" or "THC:CBD 1:1:1"
        match = re.search(pattern, product_name, re.IGNORECASE)

        if match:
            extracted_ratio = match.group(1).strip()
            # Remove ratio from product name
            clean_name = re.sub(pattern, '', product_name, flags=re.IGNORECASE).strip()
            return extracted_ratio, clean_name

        return None, product_name  # If no ratio found, return None


    def extract_clean_product_name(product_name, brand_name):
        """
        Cleans the product name by removing brand, extracted weight, and extracted ratio.
        """
        clean_name = product_name.strip()

        # ✅ Remove brand name
        if brand_name:
            brand_pattern = rf'^{re.escape(brand_name)}[\s-]*'
            clean_name = re.sub(brand_pattern, '', clean_name, flags=re.IGNORECASE).strip()

        # ✅ Extract & remove weight
        extracted_weight, clean_name = extract_weight_from_name(clean_name)

        # ✅ Extract & remove ratio
        extracted_ratio, clean_name = extract_ratio_from_name(clean_name)

        # ✅ Extract & remove quantity
        extracted_quantity, clean_name = extract_quantity_from_name(clean_name)

        # ✅ Remove extra spaces or trailing dashes
        clean_name = re.sub(r'\s*-\s*', " - ", clean_name)  # Normalize spacing around dashes
        clean_name = re.sub(r'\s+', " ", clean_name).strip()  # Remove extra spaces

        return extracted_weight, extracted_ratio, extracted_quantity, clean_name

    def extract_product_details(product_element):
        """
        Extracts price, THC, and product URL from the product element.
        
        - The product URL is built by appending the product's 'slug' to
        "https://highprofilecannabis.com/shop/martinsville-dispensary/".
        - The price is first attempted to be extracted from a variant price element.
        If not found, it falls back to an element containing a '$' sign.
        - THC is extracted from the element with class "shopitem__strain-thc".
        """
        try:
            slug = product_element.get_attribute("slug")
            if not slug:
                html = product_element.get_attribute("outerHTML")
                slug_match = re.search(r'slug="([^"]+)"', html)
                if slug_match:
                    slug = slug_match.group(1)
            url_extracted = f"https://highprofilecannabis.com/shop/martinsville-dispensary/{slug}" if slug else "N/H"
        except Exception as e:
            print("Error extracting slug")
            url_extracted = "N/H"
        
        price = None
        try:
            price_element = product_element.find_element(By.XPATH, ".//p[contains(@class, 'shopitem__listPrices-productVariants-price')]")
            price = price_element.text.strip()
        except Exception as e:
            print("Price Not Found!")
        
        if not price or price == "N/A":
            try:
                price_element = product_element.find_element(By.XPATH, ".//div[@class='shopitem__price']//div[contains(text(), '$')]")
                price = price_element.text.strip()
            except Exception as e:
                print("Fallback Price Not Found")
                price = "N/A"
        
        try:
            thc_text = product_element.find_element(By.CLASS_NAME, "shopitem__strain-thc").text.strip()
            thc = re.sub(r'^THC:\s*', '', thc_text)
        except Exception as e:
            print("No THC Found")
            thc = ""
        
        return price, thc, url_extracted


    def wait_for_page_load(driver, xpath, timeout=60):
        """Wait for a specific element to load on the page."""
        try:
            # Wait for the element to appear on the page, indicating that it is fully loaded.
            WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            print("Page loaded successfully!")
        except Exception as e:
            print(f"Error waiting for page to load: {e}")
            driver.quit()  # Close the driver after an error
            raise e
        
        
    def save_updated_excel_to_memory():
        # Create an in-memory file
        output = io.BytesIO()

        wb = load_workbook(uploaded_file)  # Load the uploaded Excel file
        wb.save(output)  # Save the workbook to the in-memory file
        output.seek(0)  # Go to the beginning of the BytesIO object

        return output

    def run_scraper(uploaded_file):
        successful = False
        try:
            driver = get_driver()  # Initialize driver
            # wait = WebDriverWait(driver, 30)  # Set the explicit wait time to 30 seconds

            # Iterate over unique categories
            unique_categories = df['Category'].dropna().unique()

            for category in unique_categories:
                st.write(f"Processing Category: {category}")
                if category not in category_url_map:
                    st.warning(f"Category {category} not found!")
                    continue
                category_url = category_url_map[category]
                print(f"🌐 Navigating to: {category_url}")
                driver.get(category_url)

                # Confirm age if prompted.
                confirm_age(driver)
                time.sleep(3)  # You can adjust this if needed

                # Wait until a specific element (like a product or category section) is loaded
                print("⏳ Waiting for the page to load...")
                # try:
                #     wait_for_page_load(driver, "//div[contains(@class, 'product-listing')]")
                # except Exception as e:
                #     st.error("Failed to load the page.")
                #     driver.quit()
                #     return

                category_rows = df[df['Category'] == category].drop_duplicates(subset=['Brand', 'Weight'])
                total_rows = len(category_rows)
                progress_bar = st.progress(0)

                for index, row in category_rows.iterrows():
                    try:
                        brand = row['Brand']
                        expected_product_name = row['Product Name']

                        # Extract weight, ratio, and clean product name using the existing function
                        extracted_weight, extracted_ratio, extracted_quantity, clean_product_name = extract_clean_product_name(expected_product_name, brand)
                        st.write(f"Processing Row {index}: Brand: {brand}, Weight: {extracted_weight}, Cleaned Name: {clean_product_name}")

                        # Update progress bar
                        progress_bar.progress((index + 1) / total_rows)

                        time.sleep(2)  # Pause to make output readable

                        weight = str(row['Weight']).strip()

                        # ✅ Convert Flower weights if needed
                        if category == "Flower" and extracted_weight and extracted_weight.endswith("g"):
                            flower_weight_map = {"7g": "1/4oz", "14g": "1/2oz", "28g": "1oz"}
                            if weight in flower_weight_map:
                                weight = flower_weight_map[extracted_weight]
                                print(f"✅ Converted Flower Weight: {extracted_weight}")

                        # ✅ Select brand
                        if not search_brand(driver, brand):
                            print(f"Brand Not Found!! Skipping row {index}...")
                            continue

                        time.sleep(2)

                        weight_applied = False
                        if weight == "SKIP_WEIGHT":
                            if extracted_weight:
                                print(f"🟡 Attempting to select weight from product name: {extracted_weight}")
                                weight_applied = select_weight(driver, extracted_weight, category, clean_product_name)
                            else:
                                print("No weight found in product name; skipping weight selection.")
                        else:
                            if category == "Accessories":
                                print("✅ Category is Accessories, selecting weight 'N/A'.")
                                weight_applied = select_weight(driver, "N/A", category, clean_product_name)
                            else:
                                weight_applied = select_weight(driver, weight, category, clean_product_name)
                                if not weight_applied and extracted_weight:
                                    print(f"🟡 Attempting extracted weight instead: {extracted_weight}")
                                    weight_applied = select_weight(driver, extracted_weight, category, clean_product_name)

                        if not weight_applied:
                            print("-------------------------------------------------------------------")
                            print(f"Weight Filter Not Pass for row {index}! Skipping product matching...")
                            clear_weight_filters(driver)  # (Optional)
                            continue

                        matches = match_product(driver, clean_product_name, brand, weight, weight_applied, extracted_quantity)

                        if matches:
                            # If weight was missing, update it in Excel
                            if weight == "SKIP_WEIGHT" and extracted_weight:
                                update_weight_in_excel(index, extracted_weight)

                            # Now handle each match
                            for single_match in matches:
                                product_name = single_match["product_name"]
                                score = single_match["score"]
                                product_elem = single_match["element"]

                                # Extract product details
                                price, thc, url_extracted = extract_product_details(product_elem)
                                st.write("------------------------------------------")
                                st.write(f"✅ Matched Product: {product_name} (Score: {score:.2f})")
                                st.write(f"✅ Extracted -> Price: {price}, THC: {thc}, URL: {url_extracted}")
                                st.write("------------------------------------------")

                                # Save extracted data to Excel
                                extract_data_and_update_excel(index, price, thc, url_extracted)

                                # Determine the correct weight to unselect
                                weight_to_unselect = weight
                                converted_weight = convert_weight_alternate(weight)
                                if weight_applied and converted_weight:
                                    weight_to_unselect = converted_weight
                                elif weight_applied and extracted_weight:
                                    weight_to_unselect = extracted_weight

                                st.write(f"🔄 Preparing to unselect weight: {weight_to_unselect}")

                        else:
                            st.warning(f"No match found for {expected_product_name}")
                    except Exception as e:
                        st.error(f"Error processing row {index}: {e}")
                        print(f"Error processing row {index}: {e}")

                driver.quit()
            successful = True

        except Exception as e:
            st.error(f"Error occurred: {e}")
            try:
                driver.quit()
            except:
                pass

        if successful:
            st.success("Scraper ran successfully!")
                        # After the scraping finishes and Excel is updated, provide the download link
            # Get the updated Excel file in memory
            updated_file = save_updated_excel_to_memory()

            st.write("Click the button below to download the updated Excel file.")
            
            # Streamlit download button
            st.download_button(
                label="Download Updated Excel File", 
                data=updated_file, 
                file_name="updated_pricing_research.xlsx", 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            driver.quit()
        else:
            st.error("Scraper failed!")

    run_scraper(uploaded_file)

